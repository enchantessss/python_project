import xlrd
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter 

wb = xlrd.open_workbook('617新增售后.xls')
sheet_names = wb.sheet_names()
print(sheet_names)

sh = wb.sheet_by_index(0)
length_row = sh.nrows
length_col = sh.ncols
print("Total Row", length_row)
print("Total Col", length_col)

deliveryID_col = 0
orderID_col = 0
phoneID_col = 0

for cx in range(sh.ncols):
    print(sh.cell_value(0, cx))
    if(sh.cell_value(0, cx) == "总订单编号"):
        print("Search in Column of orderID", cx)
        orderID_col = cx
    
    if(sh.cell_value(0, cx) == "收货人手机号"):
        print("Search in Column of phoneID", cx)
        phoneID_col = cx
    
    if(sh.cell_value(0, cx) == "物流单号"):
        print("Search in Column of deliverID", cx)
        deliveryID_col = cx


def is_contains_chinese(strs):
    for _char in strs:
        if '\u4e00' <= _char <= '\u9fa5':
            return True
    return False


def if_contain_symbol(strs):
    symbols = "~!@#$%^&amp;*()_+-*/&lt;&gt;,.[]\/"
    for symbol in symbols:
        if symbol in strs:
            return True
    else:
        return False


print("================== Start Sheet ==================")

deliveryID_list = []
orderID_list = []
phoneID_list = []
row_list =[]


for rx in range(1, sh.nrows):
    row_list.append(rx)

    deliveryID_cell = sh.cell(rx, deliveryID_col).value
    deliveryID_cell = str(deliveryID_cell).split('.')[0]

    orderID_cell = sh.cell(rx, orderID_col).value
    orderID_cell = str(orderID_cell).split('.')[0]

    phoneID_cell = sh.cell(rx, phoneID_col).value
    phoneID_cell = str(phoneID_cell).split('.')[0]

    if(orderID_cell == '' or phoneID_cell == ''):
        raise Exception("orderID or phoneID is null!!!! Please check!!")

    #print(deliveryID_cell)

    if(deliveryID_cell == ""):
        print("Exception Cell Found at Row", rx + 1) #Excel column starts from 1, while we take it start from 1 
        deliveryID_list.append('')

    elif(is_contains_chinese(deliveryID_cell)):
        print("Exception Cell Found at Row", rx + 1)
        deliveryID_list.append('')
    
    elif(if_contain_symbol(deliveryID_cell)):
        print("Exception Cell Found at Row", rx + 1)
        deliveryID_list.append('')
    
    else:
        deliveryID_list.append(deliveryID_cell)

    orderID_list.append(orderID_cell)
    phoneID_list.append(phoneID_cell)
#
##print(row_list)
##print(deliveryID_list)
##print(orderID_list)
##print(phoneID_list)



#ws: work sheet
#wb: work book
#target_wb = load_workbook('嘉丰农行小豆售后00.xlsx')
target_wb = load_workbook('蓝漂博轩补发汇总00.xlsx')
target_ws_names = target_wb.active
print(target_ws_names)
 
t_ws = target_wb.active
print(t_ws.max_row)
print(t_ws.max_column)

t_deliveryID_col = 1
t_orderID_col = 1
t_phoneID_col = 1
t_Reply_col = 1

t_deliveryID_list = []
t_orderID_list = []
t_phoneID_list = []
t_Reply_list = []

for row in range(1, 2):
    for col in range(1, int(t_ws.max_column)):
        char = get_column_letter(col)
        print(t_ws[char + str(row)].value)
        if(t_ws[char + str(row)].value == '快递单号'):
            print("Target delivery ID at column", col)
            t_deliveryID_col = col
        if(t_ws[char + str(row)].value == '手机号'):
            print("Target phone ID at column", col)
            t_phoneID_col = col
        if(t_ws[char + str(row)].value == '结果回复'):
            print("Target result Reply at column", col)
            t_Reply_col = col
        if(t_ws[char + str(row)].value == '总订单编号'):
            print("Target order ID at column", col)
            t_orderID_col = col


for row in range(1, t_ws.max_row):
        t_deliveryID_list.append(str(t_ws[get_column_letter(t_deliveryID_col) + str(row)].value))
        t_orderID_list.append(str(t_ws[get_column_letter(t_orderID_col) + str(row)].value))
        t_phoneID_list.append(str(t_ws[get_column_letter(t_phoneID_col) + str(row)].value))
        t_Reply_list.append(str(t_ws[get_column_letter(t_Reply_col) + str(row)].value))


print(">>>>>>>>>>>>>>> Target Search >>>>>>>>>>>>>>>>>>")

#print(len(deliveryID_list))
#print(len(t_deliveryID_list))

for i in range(len(deliveryID_list)):
    for j in range(len(t_deliveryID_list)):
        if(deliveryID_list[i] != ''):
            if(deliveryID_list[i] == t_deliveryID_list[j]):
                print("Found Delivery ID in target", deliveryID_list[i], t_Reply_list[j])
                
        elif(phoneID_list[i] == t_phoneID_list[j]):
            print("Found Phone ID in target", phoneID_list[i], t_Reply_list[j])
        
        elif(orderID_list[i] == t_orderID_list[j]):
            print("Found Order ID in target", orderID_list[i], t_Reply_list[j])


#target_sh = target_wb.sheet_by_index(0)
#target_length_row = target_sh.nrows
#target_length_col = target_sh.ncols
#print("Search Total Row", target_length_row)
#print("Search Total Col", target_length_col)






